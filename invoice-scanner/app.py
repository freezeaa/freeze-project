import os
import logging
from datetime import datetime

from flask import Flask, render_template, request, send_from_directory, jsonify
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
APP_HOST = os.getenv("APP_HOST", "0.0.0.0")
APP_PORT = int(os.getenv("APP_PORT", "5000"))
APP_DEBUG = os.getenv("APP_DEBUG", "false").lower() in {"1", "true", "yes", "on"}
EXPORT_DIR = os.getenv("EXPORT_DIR", os.path.join(BASE_DIR, "static", "exports"))
LOG_PATH = os.getenv("LOG_PATH", os.path.join(BASE_DIR, "scanner.log"))

os.makedirs(EXPORT_DIR, exist_ok=True)
log_dir = os.path.dirname(LOG_PATH)
if log_dir:
    os.makedirs(log_dir, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

app = Flask(__name__)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/log", methods=["POST"])
def log_scan():
    data = request.get_json(silent=True) or {}
    log.info("扫描记录 | %s", data)
    return {"ok": True}


@app.route("/api/export", methods=["POST"])
def export_excel():
    data = request.get_json(silent=True) or {}
    invoices = data.get("invoices", []) if isinstance(data, dict) else data
    employee_name = data.get("employeeName", "") if isinstance(data, dict) else ""
    if not invoices:
        return jsonify({"error": "没有数据可导出"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "发票明细"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B59FE", end_color="3B59FE", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    headers = ["发票号码", "金额(元)", "开票日期", "所属员工姓名"]
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, inv in enumerate(invoices, 2):
        ws.cell(row=row_idx, column=1, value=inv.get("number", "")).border = thin_border

        amount_cell = ws.cell(row=row_idx, column=2)
        amount_cell.border = thin_border
        amount_cell.number_format = "#,##0.00"
        try:
            amount_cell.value = float(inv.get("amount", 0))
        except (ValueError, TypeError):
            amount_cell.value = inv.get("amount", "")

        ws.cell(row=row_idx, column=3, value=inv.get("date", "")).border = thin_border
        ws.cell(row=row_idx, column=4, value=employee_name).border = thin_border

    filename = f"invoice_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    log.info("导出 Excel | 共 %d 条 | 文件: %s", len(invoices), filename)
    return jsonify({"url": f"/static/exports/{filename}"})


if __name__ == "__main__":
    print("\n" + "=" * 50)
    print("  发票扫码助手已启动")
    print(f"  监听地址: http://{APP_HOST}:{APP_PORT}")
    print("=" * 50 + "\n")
    app.run(host=APP_HOST, port=APP_PORT, debug=APP_DEBUG)
